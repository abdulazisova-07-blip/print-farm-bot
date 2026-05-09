import asyncio
import logging
from datetime import datetime
from aiogram import Bot, Dispatcher, F
from aiogram.filters import Command
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
import aiosqlite
import openpyxl
from io import BytesIO

import os
import threading
from flask import Flask

app = Flask(__name__)

@app.route('/')
def home():
    return "Bot is running"

def run_flask():
    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port)

# ---------- НАСТРОЙКИ ----------
TOKEN = "8785273956:AAF8mdNuhjeM2Onrbqs0xeG3fYG-arQNI9k"
ARTEM_ID = 1172985519
BAZA_ID = 987654321
ADMIN_IDS = {ARTEM_ID, BAZA_ID}
DB_NAME = "farm.db"

# ---------- ИНИЦИАЛИЗАЦИЯ БАЗЫ ДАННЫХ ----------
async def init_db():
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("""
            CREATE TABLE IF NOT EXISTS printers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ip TEXT UNIQUE NOT NULL,
                status TEXT DEFAULT 'работает',
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS breakdowns (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                printer_id INTEGER NOT NULL,
                type TEXT NOT NULL DEFAULT 'broken',
                reason TEXT,
                reported_by INTEGER,
                reported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                resolved_at TIMESTAMP,
                resolved_by INTEGER,
                FOREIGN KEY(printer_id) REFERENCES printers(id)
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS part_codes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT UNIQUE NOT NULL,
                name TEXT NOT NULL,
                photo_file_id TEXT
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS boxes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT NOT NULL,
                name TEXT,
                quantity INTEGER NOT NULL,
                user_id INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS shipments (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                code TEXT,
                name TEXT,
                quantity INTEGER,
                user_id INTEGER,
                shipped_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.execute("""
            CREATE TABLE IF NOT EXISTS operators (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                telegram_id INTEGER UNIQUE NOT NULL,
                username TEXT,
                full_name TEXT,
                added_by INTEGER,
                added_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)
        await db.commit()

# ---------- СОСТОЯНИЯ FSM ----------
class PrinterBreak(StatesGroup):
    waiting_for_ip = State()
    waiting_for_reason = State()

class PrinterDefect(StatesGroup):
    waiting_for_ip = State()
    waiting_for_reason = State()

class PartAdd(StatesGroup):
    waiting_for_selection = State()
    waiting_for_qty = State()
    confirm = State()

class AdminAddPrinter(StatesGroup):
    waiting_for_ip = State()

class AdminRemovePrinter(StatesGroup):
    waiting_for_ip = State()

class AdminStatHistory(StatesGroup):
    waiting_for_ip = State()

class AdminPartCodeAdd(StatesGroup):
    waiting_for_code = State()
    waiting_for_name = State()
    waiting_for_photo = State()

class AdminPartCodeDelete(StatesGroup):
    waiting_for_selection = State()
    confirm = State()

class ShipmentSelect(StatesGroup):
    collecting = State()
    confirm = State()

class ReportOther(StatesGroup):
    waiting_for_text = State()

class AddOperator(StatesGroup):
    waiting_for_id = State()

class RemoveOperator(StatesGroup):
    waiting_for_selection = State()
    confirm = State()

# ---------- МЕНЮ (ТРЁХУРОВНЕВОЕ) ----------
def main_menu(user_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="🖨️ Принтеры", callback_data="menu_printers")
    kb.button(text="📦 Детали", callback_data="menu_parts_main")
    kb.button(text="📞 Связь", callback_data="menu_comm")
    if user_id in ADMIN_IDS:
        kb.button(text="👥 Сотрудники", callback_data="menu_staff")
    kb.adjust(2)
    return kb.as_markup()

def printers_menu(user_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="Сломался принтер", callback_data="menu_break")
    kb.button(text="Брак принтера", callback_data="menu_defect")
    kb.button(text="Вернуть в работу", callback_data="menu_return")
    kb.button(text="Список неработающих", callback_data="menu_list_broken")
    kb.button(text="История принтера", callback_data="menu_stat_history_op")
    if user_id in ADMIN_IDS:
        kb.button(text="➕ Добавить принтер", callback_data="menu_add_printer")
        kb.button(text="🗑️ Снять принтер", callback_data="menu_remove_printer")
        kb.button(text="📊 Статистика", callback_data="menu_stats")
    kb.button(text="🔙 Главное меню", callback_data="back_to_main")
    kb.adjust(2)
    return kb.as_markup()

def parts_main_menu(user_id: int):
    kb = InlineKeyboardBuilder()
    kb.button(text="Добавить коробку", callback_data="menu_add_part")
    if user_id in ADMIN_IDS:
        kb.button(text="📋 Склад (группировка)", callback_data="menu_parts_grouped")
    else:
        kb.button(text="📋 Склад", callback_data="menu_parts_list")
    kb.button(text="🚚 Отгрузка", callback_data="menu_ship")
    if user_id in ADMIN_IDS:
        kb.button(text="📚 Справочник деталей", callback_data="menu_part_codes")
        kb.button(text="📥 Выгрузить Excel", callback_data="menu_export")
    kb.button(text="🔙 Главное меню", callback_data="back_to_main")
    kb.adjust(1)
    return kb.as_markup()

def comm_menu():
    kb = InlineKeyboardBuilder()
    kb.button(text="📞 Сообщить начальству", callback_data="menu_report")
    kb.button(text="🔙 Главное меню", callback_data="back_to_main")
    return kb.as_markup()

def staff_menu():
    kb = InlineKeyboardBuilder()
    kb.button(text="➕ Добавить оператора", callback_data="staff_add")
    kb.button(text="🗑️ Удалить оператора", callback_data="staff_remove")
    kb.button(text="📋 Список операторов", callback_data="staff_list")
    kb.button(text="🔙 Главное меню", callback_data="back_to_main")
    return kb.as_markup()

# ---------- ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ----------
async def is_admin(user_id: int) -> bool:
    return user_id in ADMIN_IDS

async def is_operator(user_id: int) -> bool:
    if user_id in ADMIN_IDS:
        return True
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT 1 FROM operators WHERE telegram_id = ?", (user_id,))
        return await cursor.fetchone() is not None

async def get_printer_id_by_ip(ip: str) -> int | None:
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT id FROM printers WHERE ip = ? AND status != 'снят'", (ip,))
        row = await cursor.fetchone()
        return row[0] if row else None

async def add_printer_to_db(ip: str) -> bool:
    async with aiosqlite.connect(DB_NAME) as db:
        try:
            await db.execute("INSERT INTO printers (ip, status) VALUES (?, 'работает')", (ip,))
            await db.commit()
            return True
        except aiosqlite.IntegrityError:
            return False

async def get_available_part_codes():
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT code, name, photo_file_id FROM part_codes ORDER BY name")
        return await cursor.fetchall()

# ---------- ПРОВЕРКА ДОСТУПА ПРИ /start ----------
async def start_command(message: Message):
    user_id = message.from_user.id
    if not await is_operator(user_id):
        await message.answer("⛔ Доступ запрещён. Обратитесь к администратору для добавления в список сотрудников.")
        return
    await message.answer("👋 Добро пожаловать в систему управления фермой!",
                         reply_markup=main_menu(user_id))

# ---------- ОБРАБОТЧИКИ ГЛАВНОГО МЕНЮ ----------
async def back_to_main(callback: CallbackQuery):
    await callback.message.edit_text("Главное меню:", reply_markup=main_menu(callback.from_user.id))
    await callback.answer()

async def menu_printers(callback: CallbackQuery):
    await callback.message.edit_text("🖨️ Принтеры:", reply_markup=printers_menu(callback.from_user.id))
    await callback.answer()

async def menu_parts_main(callback: CallbackQuery):
    await callback.message.edit_text("📦 Детали:", reply_markup=parts_main_menu(callback.from_user.id))
    await callback.answer()

async def menu_comm(callback: CallbackQuery):
    await callback.message.edit_text("📞 Связь:", reply_markup=comm_menu())
    await callback.answer()

async def menu_staff(callback: CallbackQuery):
    await callback.message.edit_text("👥 Управление сотрудниками:", reply_markup=staff_menu())
    await callback.answer()

# ---------- ПРИНТЕРЫ: СЛОМАЛСЯ / БРАК / ВЕРНУТЬ / СПИСОК / ИСТОРИЯ ----------
async def menu_break(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введите IP-адрес сломавшегося принтера:",
                                     reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="menu_printers").as_markup())
    await state.set_state(PrinterBreak.waiting_for_ip)
    await callback.answer()

async def break_ip(message: Message, state: FSMContext):
    ip = message.text.strip()
    printer_id = await get_printer_id_by_ip(ip)
    if not printer_id:
        await message.answer("❌ Принтер с таким IP не найден.", reply_markup=printers_menu(message.from_user.id))
        await state.clear()
        return
    await state.update_data(printer_id=printer_id, ip=ip)
    await message.answer("Опишите причину поломки:", reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="menu_printers").as_markup())
    await state.set_state(PrinterBreak.waiting_for_reason)

async def break_reason(message: Message, state: FSMContext):
    reason = message.text.strip()
    data = await state.get_data()
    printer_id = data["printer_id"]
    ip = data["ip"]
    user_id = message.from_user.id

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("INSERT INTO breakdowns (printer_id, type, reason, reported_by) VALUES (?, 'broken', ?, ?)",
                         (printer_id, reason, user_id))
        await db.execute("UPDATE printers SET status = 'сломан' WHERE id = ?", (printer_id,))
        await db.commit()

    await message.answer(f"✅ Принтер {ip} отмечен как сломаный.",
                         reply_markup=main_menu(user_id))
    await state.clear()

async def menu_defect(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введите IP-адрес принтера, дающего брак:",
                                     reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="menu_printers").as_markup())
    await state.set_state(PrinterDefect.waiting_for_ip)
    await callback.answer()

async def defect_ip(message: Message, state: FSMContext):
    ip = message.text.strip()
    printer_id = await get_printer_id_by_ip(ip)
    if not printer_id:
        await message.answer("❌ Принтер с таким IP не найден.", reply_markup=printers_menu(message.from_user.id))
        await state.clear()
        return
    await state.update_data(printer_id=printer_id, ip=ip)
    await message.answer("Опишите характер брака:", reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="menu_printers").as_markup())
    await state.set_state(PrinterDefect.waiting_for_reason)

async def defect_reason(message: Message, state: FSMContext):
    reason = message.text.strip()
    data = await state.get_data()
    printer_id = data["printer_id"]
    ip = data["ip"]
    user_id = message.from_user.id

    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("INSERT INTO breakdowns (printer_id, type, reason, reported_by) VALUES (?, 'defect', ?, ?)",
                         (printer_id, reason, user_id))
        await db.execute("UPDATE printers SET status = 'сломан' WHERE id = ?", (printer_id,))
        await db.commit()

    await message.answer(f"⚠️ Брак зафиксирован на принтере {ip}.",
                         reply_markup=main_menu(user_id))
    await state.clear()

async def menu_return(callback: CallbackQuery):
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("""
            SELECT p.ip, b.id, b.type, b.reason, b.reported_at
            FROM printers p
            JOIN breakdowns b ON p.id = b.printer_id
            WHERE p.status = 'сломан' AND b.resolved_at IS NULL
            ORDER BY b.reported_at DESC
        """)
        rows = await cursor.fetchall()

    if not rows:
        await callback.message.edit_text("✅ Все принтеры работают!", reply_markup=printers_menu(callback.from_user.id))
        await callback.answer()
        return

    kb = InlineKeyboardBuilder()
    for ip, bid, typ, reason, reported_at in rows:
        prefix = "⚠️" if typ == "defect" else "🖨️"
        date_str = datetime.strptime(reported_at, "%Y-%m-%d %H:%M:%S").strftime("%d.%m %H:%M") if reported_at else "?"
        kb.button(text=f"{prefix} {ip} — {reason} ({date_str})", callback_data=f"return_{bid}")
    kb.button(text="🔙 Назад", callback_data="menu_printers")
    kb.adjust(1)
    await callback.message.edit_text("Выберите принтер, который починили:", reply_markup=kb.as_markup())
    await callback.answer()

async def return_printer(callback: CallbackQuery):
    breakdown_id = int(callback.data.split("_")[1])
    user_id = callback.from_user.id

    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT printer_id FROM breakdowns WHERE id = ?", (breakdown_id,))
        row = await cursor.fetchone()
        if not row:
            await callback.answer("Ошибка: запись не найдена.", show_alert=True)
            return
        printer_id = row[0]
        await db.execute("UPDATE breakdowns SET resolved_at = CURRENT_TIMESTAMP, resolved_by = ? WHERE id = ?",
                         (user_id, breakdown_id))
        await db.execute("UPDATE printers SET status = 'работает' WHERE id = ?", (printer_id,))
        cursor = await db.execute("SELECT ip FROM printers WHERE id = ?", (printer_id,))
        ip = (await cursor.fetchone())[0]
        await db.commit()

    await callback.message.edit_text(f"✅ Принтер {ip} возвращён в работу.",
                                     reply_markup=printers_menu(user_id))
    await callback.answer()

async def menu_list_broken(callback: CallbackQuery):
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("""
            SELECT p.ip, b.type, b.reason, b.reported_at
            FROM printers p
            JOIN breakdowns b ON p.id = b.printer_id
            WHERE p.status = 'сломан' AND b.resolved_at IS NULL
            ORDER BY b.reported_at DESC
        """)
        rows = await cursor.fetchall()
    if not rows:
        await callback.message.edit_text("✅ Все принтеры работают.", reply_markup=printers_menu(callback.from_user.id))
        await callback.answer()
        return
    text = "🟥 Неработающие принтеры:\n"
    for ip, typ, reason, reported_at in rows:
        prefix = "⚠️" if typ == "defect" else "🖨️"
        date_str = datetime.strptime(reported_at, "%Y-%m-%d %H:%M:%S").strftime("%d.%m %H:%M")
        text += f"{prefix} {ip} — {reason} (с {date_str})\n"
    kb = InlineKeyboardBuilder()
    kb.button(text="🔙 Назад", callback_data="menu_printers")
    await callback.message.edit_text(text, reply_markup=kb.as_markup())
    await callback.answer()

async def menu_stat_history_op(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введите IP-адрес принтера:",
                                     reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="menu_printers").as_markup())
    await state.set_state(AdminStatHistory.waiting_for_ip)
    await callback.answer()

async def stat_history_show_op(message: Message, state: FSMContext):
    ip = message.text.strip()
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("""
            SELECT b.type, b.reason, b.reported_at, b.resolved_at
            FROM breakdowns b
            JOIN printers p ON b.printer_id = p.id
            WHERE p.ip = ? AND p.status != 'снят'
            ORDER BY b.reported_at DESC
        """, (ip,))
        rows = await cursor.fetchall()
    if not rows:
        await message.answer(f"Принтер {ip} не найден или нет истории.",
                             reply_markup=main_menu(message.from_user.id))
    else:
        text = f"История принтера {ip}:\n\n"
        for typ, reason, reported, resolved in rows:
            prefix = "⚠️ Брак" if typ == "defect" else "🖨️ Поломка"
            rep_date = reported[:16] if reported else "?"
            res_date = resolved[:16] if resolved else "не починен"
            text += f"{prefix}: {reason}\n  {rep_date} — починен: {res_date}\n\n"
        await message.answer(text, reply_markup=main_menu(message.from_user.id))
    await state.clear()

# ---------- ДЕТАЛИ: ДОБАВЛЕНИЕ КОРОБКИ, СКЛАД, ОТГРУЗКА, СПРАВОЧНИК ----------
async def menu_add_part(callback: CallbackQuery, state: FSMContext):
    parts = await get_available_part_codes()
    if not parts:
        await callback.message.edit_text("❌ Справочник деталей пуст. Обратитесь к администратору.",
                                         reply_markup=main_menu(callback.from_user.id))
        await callback.answer()
        return
    kb = InlineKeyboardBuilder()
    for code, name, _ in parts:
        kb.button(text=name, callback_data=f"selectpart_{code}")
    kb.button(text="🔙 Назад", callback_data="menu_parts_main")
    kb.adjust(1)
    await callback.message.edit_text("Выберите деталь из справочника:", reply_markup=kb.as_markup())
    await state.set_state(PartAdd.waiting_for_selection)
    await callback.answer()

async def part_selected(callback: CallbackQuery, state: FSMContext):
    code = callback.data.split("_", 1)[1]
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT name, photo_file_id FROM part_codes WHERE code = ?", (code,))
        row = await cursor.fetchone()
        if not row:
            await callback.answer("Деталь не найдена.", show_alert=True)
            return
        name, photo = row
    await state.update_data(code=code, name=name)
    text = f"🔹 Код: {code}\n🔹 Название: {name}\n\nВведите количество штук в коробке:"
    kb = InlineKeyboardBuilder()
    kb.button(text="↩️ Выбрать другую", callback_data="menu_add_part")
    kb.button(text="🔙 Назад", callback_data="menu_parts_main")
    await callback.message.edit_text(text, reply_markup=kb.as_markup())
    if photo:
        await callback.message.answer_photo(photo, caption="Фото детали")
    await state.set_state(PartAdd.waiting_for_qty)
    await callback.answer()

async def part_qty(message: Message, state: FSMContext):
    try:
        qty = int(message.text)
        if qty <= 0:
            raise ValueError
    except ValueError:
        await message.answer("❌ Введите целое положительное число.")
        return
    data = await state.get_data()
    await state.update_data(qty=qty)
    text = f"🔹 Код: {data['code']}\n🔹 Название: {data['name']}\n🔹 Количество: {qty} шт.\n\nВсё правильно?"
    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Сохранить", callback_data="confirm_add_part")
    kb.button(text="↩️ Выбрать другую", callback_data="menu_add_part")
    kb.button(text="🔙 Отмена", callback_data="menu_parts_main")
    await message.answer(text, reply_markup=kb.as_markup())
    await state.set_state(PartAdd.confirm)

async def confirm_add_part(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    code, name, qty = data["code"], data["name"], data["qty"]
    user_id = callback.from_user.id
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("INSERT INTO boxes (code, name, quantity, user_id) VALUES (?, ?, ?, ?)",
                         (code, name, qty, user_id))
        await db.commit()
    await callback.message.edit_text(f"✅ Коробка сохранена: {code} — {name}, {qty} шт.",
                                     reply_markup=main_menu(user_id))
    await state.clear()
    await callback.answer()

async def menu_parts_list(callback: CallbackQuery):
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT code, name, quantity, created_at FROM boxes ORDER BY created_at DESC")
        rows = await cursor.fetchall()
    if not rows:
        await callback.message.edit_text("📦 Склад пуст.", reply_markup=main_menu(callback.from_user.id))
        await callback.answer()
        return
    text = "📋 Список коробок:\n\n"
    for code, name, qty, created in rows:
        date_str = created[:16] if created else "?"
        text += f"{code} — {name}, {qty} шт. ({date_str})\n"
    await callback.message.edit_text(text, reply_markup=main_menu(callback.from_user.id))
    await callback.answer()

async def menu_parts_grouped(callback: CallbackQuery):
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("""
            SELECT code, name, SUM(quantity) as total, GROUP_CONCAT(quantity, '+') as details, COUNT(*) as cnt
            FROM boxes GROUP BY code, name ORDER BY code
        """)
        rows = await cursor.fetchall()
    if not rows:
        await callback.message.edit_text("📦 Склад пуст.", reply_markup=main_menu(callback.from_user.id))
        await callback.answer()
        return
    text = "📊 Склад (группировка):\n\n"
    for code, name, total, details, cnt in rows:
        text += f"{code} — {name}: {total} шт. ({cnt} короб.: {details})\n"
    kb = InlineKeyboardBuilder()
    kb.button(text="📋 Показать списком", callback_data="menu_parts_list")
    kb.button(text="🔙 Назад", callback_data="menu_parts_main")
    await callback.message.edit_text(text, reply_markup=kb.as_markup())
    await callback.answer()

# ---------- ОТГРУЗКА (МНОЖЕСТВЕННЫЙ ВЫБОР КОРОБОК) ----------
async def menu_ship(callback: CallbackQuery, state: FSMContext):
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT id, code, name, quantity FROM boxes ORDER BY created_at DESC")
        boxes = await cursor.fetchall()
    if not boxes:
        await callback.message.edit_text("🚚 Нет коробок для отгрузки.", reply_markup=main_menu(callback.from_user.id))
        await callback.answer()
        return
    kb = InlineKeyboardBuilder()
    for box_id, code, name, qty in boxes:
        kb.button(text=f"{code} — {name} ({qty} шт.)", callback_data=f"shipsel_{box_id}")
    kb.button(text="✅ Отгрузить выбранные", callback_data="ship_confirm")
    kb.button(text="🔙 Назад", callback_data="menu_parts_main")
    kb.adjust(1)
    await callback.message.edit_text("Выберите коробки для отгрузки (можно несколько):", reply_markup=kb.as_markup())
    await state.update_data(selected_boxes=set())
    await state.set_state(ShipmentSelect.collecting)
    await callback.answer()

async def ship_sel_toggle(callback: CallbackQuery, state: FSMContext):
    box_id = int(callback.data.split("_")[1])
    data = await state.get_data()
    selected = data.get("selected_boxes", set())
    if box_id in selected:
        selected.discard(box_id)
    else:
        selected.add(box_id)
    await state.update_data(selected_boxes=selected)
    await callback.answer(f"Выбрано: {len(selected)} коробок")

async def ship_confirm_handler(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    selected = data.get("selected_boxes", set())
    if not selected:
        await callback.answer("Не выбрано ни одной коробки.", show_alert=True)
        return
    user_id = callback.from_user.id
    async with aiosqlite.connect(DB_NAME) as db:
        for box_id in selected:
            cursor = await db.execute("SELECT code, name, quantity FROM boxes WHERE id = ?", (box_id,))
            box = await cursor.fetchone()
            if box:
                code, name, qty = box
                await db.execute("INSERT INTO shipments (code, name, quantity, user_id) VALUES (?, ?, ?, ?)",
                                 (code, name, qty, user_id))
                await db.execute("DELETE FROM boxes WHERE id = ?", (box_id,))
        await db.commit()
    await callback.message.edit_text(f"✅ Отгружено коробок: {len(selected)}",
                                     reply_markup=main_menu(user_id))
    await state.clear()
    await callback.answer()

# ---------- СПРАВОЧНИК ДЕТАЛЕЙ (АДМИНЫ) ----------
async def menu_part_codes(callback: CallbackQuery):
    kb = InlineKeyboardBuilder()
    kb.button(text="➕ Добавить деталь", callback_data="partcode_add")
    kb.button(text="🗑️ Удалить деталь", callback_data="partcode_delete")
    kb.button(text="📋 Показать все", callback_data="partcode_list")
    kb.button(text="🔙 Назад", callback_data="back_to_main")
    await callback.message.edit_text("📚 Справочник деталей:", reply_markup=kb.as_markup())
    await callback.answer()

async def partcode_add_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введите код новой детали:",
                                     reply_markup=InlineKeyboardBuilder().button(text="🔙 Отмена", callback_data="menu_part_codes").as_markup())
    await state.set_state(AdminPartCodeAdd.waiting_for_code)
    await callback.answer()

async def partcode_add_code(message: Message, state: FSMContext):
    await state.update_data(code=message.text.strip().upper())
    await message.answer("Введите название детали:",
                         reply_markup=InlineKeyboardBuilder().button(text="🔙 Отмена", callback_data="menu_part_codes").as_markup())
    await state.set_state(AdminPartCodeAdd.waiting_for_name)

async def partcode_add_name(message: Message, state: FSMContext):
    await state.update_data(name=message.text.strip())
    await message.answer("Пришлите фото детали (или нажмите кнопку \"Пропустить\"):",
                         reply_markup=InlineKeyboardBuilder()
                         .button(text="Пропустить", callback_data="partcode_skip_photo")
                         .button(text="🔙 Отмена", callback_data="menu_part_codes").as_markup())
    await state.set_state(AdminPartCodeAdd.waiting_for_photo)

async def partcode_add_photo(message: Message, state: FSMContext):
    photo_id = message.photo[-1].file_id if message.photo else None
    data = await state.get_data()
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("INSERT INTO part_codes (code, name, photo_file_id) VALUES (?, ?, ?)",
                         (data["code"], data["name"], photo_id))
        await db.commit()
    await message.answer("✅ Деталь добавлена в справочник.", reply_markup=main_menu(message.from_user.id))
    await state.clear()

async def partcode_skip_photo(callback: CallbackQuery, state: FSMContext):
    data = await state.get_data()
    async with aiosqlite.connect(DB_NAME) as db:
        await db.execute("INSERT INTO part_codes (code, name, photo_file_id) VALUES (?, ?, NULL)",
                         (data["code"], data["name"]))
        await db.commit()
    await callback.message.edit_text("✅ Деталь добавлена без фото.", reply_markup=main_menu(callback.from_user.id))
    await state.clear()
    await callback.answer()

# ---------- УПРАВЛЕНИЕ ОПЕРАТОРАМИ ----------
async def staff_add_start(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("Введите Telegram ID нового оператора (цифры):",
                                     reply_markup=InlineKeyboardBuilder().button(text="🔙 Отмена", callback_data="menu_staff").as_markup())
    await state.set_state(AddOperator.waiting_for_id)
    await callback.answer()

async def staff_add_id(message: Message, state: FSMContext):
    try:
        tg_id = int(message.text.strip())
    except ValueError:
        await message.answer("❌ Введите числовой ID.")
        return
    async with aiosqlite.connect(DB_NAME) as db:
        try:
            await db.execute("INSERT INTO operators (telegram_id, added_by) VALUES (?, ?)",
                             (tg_id, message.from_user.id))
            await db.commit()
        except aiosqlite.IntegrityError:
            await message.answer("❌ Оператор с таким ID уже существует.", reply_markup=main_menu(message.from_user.id))
            await state.clear()
            return
    await message.answer("✅ Оператор добавлен.", reply_markup=main_menu(message.from_user.id))
    await state.clear()

async def staff_list(callback: CallbackQuery):
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT telegram_id, username, full_name, added_at FROM operators ORDER BY added_at")
        rows = await cursor.fetchall()
    if not rows:
        await callback.message.edit_text("Список операторов пуст.", reply_markup=staff_menu())
        await callback.answer()
        return
    text = "👥 Список операторов:\n\n"
    for tid, uname, fname, added in rows:
        text += f"ID: {tid} | @{uname or 'нет'} | {fname or '—'} (добавлен {added[:10]})\n"
    kb = InlineKeyboardBuilder()
    kb.button(text="🔙 Назад", callback_data="menu_staff")
    await callback.message.edit_text(text, reply_markup=kb.as_markup())
    await callback.answer()

# ---------- ВЫГРУЗКА EXCEL (АДМИНЫ) ----------
async def menu_export(callback: CallbackQuery):
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "Принтеры"
    ws1.append(["IP", "Статус", "Дата добавления"])
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT ip, status, added_at FROM printers ORDER BY id")
        for row in await cursor.fetchall():
            ws1.append(list(row))
    ws2 = wb.create_sheet("Поломки и брак")
    ws2.append(["IP", "Тип", "Причина", "Дата", "Починен"])
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("""
            SELECT p.ip, b.type, b.reason, b.reported_at, b.resolved_at
            FROM breakdowns b JOIN printers p ON b.printer_id = p.id ORDER BY b.reported_at DESC
        """)
        for row in await cursor.fetchall():
            ws2.append(list(row))
    ws3 = wb.create_sheet("Склад")
    ws3.append(["Код", "Название", "Количество", "Дата упаковки", "Упаковал"])
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT code, name, quantity, created_at, user_id FROM boxes ORDER BY created_at DESC")
        for row in await cursor.fetchall():
            ws3.append(list(row))
    ws4 = wb.create_sheet("Отгрузки")
    ws4.append(["Код", "Название", "Количество", "Дата отгрузки", "Отгрузил"])
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT code, name, quantity, shipped_at, user_id FROM shipments ORDER BY shipped_at DESC")
        for row in await cursor.fetchall():
            ws4.append(list(row))

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    await callback.message.answer_document(
        FSInputFile(bio, filename="farm_report.xlsx"),
        caption="📊 Ежедневный отчёт",
        reply_markup=main_menu(callback.from_user.id)
    )
    await callback.answer()

# ---------- БЭКАП БАЗЫ ПО КНОПКЕ (АДМИНЫ) ----------
async def menu_backup(callback: CallbackQuery):
    await callback.message.answer_document(
        FSInputFile(DB_NAME, filename=f"backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.db"),
        caption="🗄️ Бэкап базы данных",
        reply_markup=main_menu(callback.from_user.id)
    )
    await callback.answer()

# ---------- АДМИНКА: ДОБАВЛЕНИЕ ПРИНТЕРА ----------
async def menu_add_printer(callback: CallbackQuery, state: FSMContext):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Доступ запрещён", show_alert=True)
        return
    await callback.message.edit_text("➕ Введите IP-адрес нового принтера:",
                                     reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="menu_printers").as_markup())
    await state.set_state(AdminAddPrinter.waiting_for_ip)
    await callback.answer()

async def add_printer_ip(message: Message, state: FSMContext):
    ip = message.text.strip()
    success = await add_printer_to_db(ip)
    if success:
        await message.answer(f"✅ Принтер {ip} добавлен в реестр.", reply_markup=main_menu(message.from_user.id))
    else:
        await message.answer(f"❌ Принтер с IP {ip} уже существует.", reply_markup=printers_menu(message.from_user.id))
    await state.clear()

# ---------- АДМИНКА: СНЯТИЕ ПРИНТЕРА ----------
async def menu_remove_printer(callback: CallbackQuery, state: FSMContext):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Доступ запрещён", show_alert=True)
        return
    await callback.message.edit_text("🗑️ Введите IP-адрес принтера для снятия с производства:",
                                     reply_markup=InlineKeyboardBuilder().button(text="🔙 Назад", callback_data="menu_printers").as_markup())
    await state.set_state(AdminRemovePrinter.waiting_for_ip)
    await callback.answer()

async def remove_printer_ip(message: Message, state: FSMContext):
    ip = message.text.strip()
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT id FROM printers WHERE ip = ? AND status != 'снят'", (ip,))
        if not await cursor.fetchone():
            await message.answer("❌ Принтер не найден.", reply_markup=printers_menu(message.from_user.id))
        else:
            await db.execute("UPDATE printers SET status = 'снят' WHERE ip = ?", (ip,))
            await db.commit()
            await message.answer(f"✅ Принтер {ip} снят с производства.", reply_markup=main_menu(message.from_user.id))
    await state.clear()

# ---------- АДМИНКА: СТАТИСТИКА ----------
async def menu_stats(callback: CallbackQuery):
    if not await is_admin(callback.from_user.id):
        await callback.answer("⛔ Доступ запрещён", show_alert=True)
        return
    async with aiosqlite.connect(DB_NAME) as db:
        cursor = await db.execute("SELECT COUNT(*) FROM printers WHERE status = 'сломан'")
        broken_now = (await cursor.fetchone())[0]
        cursor = await db.execute("""
            SELECT p.ip, COUNT(b.id) as cnt
            FROM breakdowns b
            JOIN printers p ON b.printer_id = p.id
            GROUP BY p.id
            ORDER BY cnt DESC
            LIMIT 5
        """)
        top = await cursor.fetchall()

    text = f"📊 Сейчас сломано: {broken_now}\n\n"
    if top:
        text += "🔝 Топ-5 проблемных принтеров:\n"
        for i, (ip, cnt) in enumerate(top, 1):
            text += f"{i}. {ip} — {cnt} поломок\n"
    else:
        text += "Нет данных о поломках."

    kb = InlineKeyboardBuilder()
    kb.button(text="📜 История по IP", callback_data="menu_stat_history_op")
    kb.button(text="🔙 Назад", callback_data="menu_printers")
    await callback.message.edit_text(text, reply_markup=kb.as_markup())
    await callback.answer()

# ---------- КНОПКА "СООБЩИТЬ НАЧАЛЬСТВУ" ----------
async def menu_report(callback: CallbackQuery):
    kb = InlineKeyboardBuilder()
    kb.button(text="📦 Закончились коробки", callback_data="report_boxes")
    kb.button(text="🧤 Нужны перчатки", callback_data="report_gloves")
    kb.button(text="🛠️ Другое", callback_data="report_other")
    kb.button(text="🔙 Назад", callback_data="menu_comm")
    await callback.message.edit_text("Выберите тип сообщения:", reply_markup=kb.as_markup())
    await callback.answer()

async def report_boxes(callback: CallbackQuery):
    for admin_id in ADMIN_IDS:
        await callback.bot.send_message(admin_id,
            f"📩 Сообщение от оператора @{callback.from_user.username or callback.from_user.id}:\nЗакончились коробки!")
    await callback.message.edit_text("✅ Сообщение отправлено начальству.",
                                     reply_markup=comm_menu())
    await callback.answer()

async def report_gloves(callback: CallbackQuery):
    for admin_id in ADMIN_IDS:
        await callback.bot.send_message(admin_id,
            f"📩 Сообщение от оператора @{callback.from_user.username or callback.from_user.id}:\nНужны перчатки!")
    await callback.message.edit_text("✅ Сообщение отправлено.",
                                     reply_markup=comm_menu())
    await callback.answer()

async def report_other(callback: CallbackQuery, state: FSMContext):
    await callback.message.edit_text("✏️ Введите текст сообщения:")
    await state.set_state(ReportOther.waiting_for_text)
    await callback.answer()

async def report_other_text(message: Message, state: FSMContext):
    text = message.text.strip()
    for admin_id in ADMIN_IDS:
        await message.bot.send_message(admin_id,
            f"📩 От оператора @{message.from_user.username or message.from_user.id}:\n{text}")
    await message.answer("✅ Сообщение отправлено начальству.",
                         reply_markup=comm_menu())
    await state.clear()

# ---------- ЗАПУСК ----------
async def main():
    logging.basicConfig(level=logging.INFO)
    await init_db()

    bot = Bot(token=TOKEN)
    dp = Dispatcher()

    # Старт и проверка доступа
    dp.message.register(start_command, Command("start"))

    # Обработчики главного меню
    dp.callback_query.register(back_to_main, F.data == "back_to_main")
    dp.callback_query.register(menu_printers, F.data == "menu_printers")
    dp.callback_query.register(menu_parts_main, F.data == "menu_parts_main")
    dp.callback_query.register(menu_comm, F.data == "menu_comm")
    dp.callback_query.register(menu_staff, F.data == "menu_staff")

    # Принтеры
    dp.callback_query.register(menu_break, F.data == "menu_break")
    dp.callback_query.register(menu_defect, F.data == "menu_defect")
    dp.callback_query.register(menu_return, F.data == "menu_return")
    dp.callback_query.register(return_printer, F.data.startswith("return_"))
    dp.callback_query.register(menu_list_broken, F.data == "menu_list_broken")
    dp.callback_query.register(menu_stat_history_op, F.data == "menu_stat_history_op")

    # Детали
    dp.callback_query.register(menu_add_part, F.data == "menu_add_part")
    dp.callback_query.register(part_selected, F.data.startswith("selectpart_"))
    dp.callback_query.register(confirm_add_part, F.data == "confirm_add_part")
    dp.callback_query.register(menu_parts_list, F.data == "menu_parts_list")
    dp.callback_query.register(menu_parts_grouped, F.data == "menu_parts_grouped")
    dp.callback_query.register(menu_ship, F.data == "menu_ship")
    dp.callback_query.register(ship_sel_toggle, F.data.startswith("shipsel_"))
    dp.callback_query.register(ship_confirm_handler, F.data == "ship_confirm")

    # Справочник деталей
    dp.callback_query.register(menu_part_codes, F.data == "menu_part_codes")
    dp.callback_query.register(partcode_add_start, F.data == "partcode_add")
    dp.callback_query.register(partcode_skip_photo, F.data == "partcode_skip_photo")

    # Отгрузка
    dp.callback_query.register(ship_sel_toggle, F.data.startswith("shipsel_"))
    dp.callback_query.register(ship_confirm_handler, F.data == "ship_confirm")

    # Связь
    dp.callback_query.register(menu_report, F.data == "menu_report")
    dp.callback_query.register(report_boxes, F.data == "report_boxes")
    dp.callback_query.register(report_gloves, F.data == "report_gloves")
    dp.callback_query.register(report_other, F.data == "report_other")

    # Админские функции
    dp.callback_query.register(menu_stats, F.data == "menu_stats")
    dp.callback_query.register(menu_add_printer, F.data == "menu_add_printer")
    dp.callback_query.register(menu_remove_printer, F.data == "menu_remove_printer")
    dp.callback_query.register(menu_export, F.data == "menu_export")
    dp.callback_query.register(menu_backup, F.data == "menu_backup")  # кнопка бэкапа

    # Управление операторами
    dp.callback_query.register(staff_add_start, F.data == "staff_add")
    dp.callback_query.register(staff_list, F.data == "staff_list")

    # Состояния
    dp.message.register(break_ip, PrinterBreak.waiting_for_ip)
    dp.message.register(break_reason, PrinterBreak.waiting_for_reason)
    dp.message.register(defect_ip, PrinterDefect.waiting_for_ip)
    dp.message.register(defect_reason, PrinterDefect.waiting_for_reason)
    dp.message.register(stat_history_show_op, AdminStatHistory.waiting_for_ip)
    dp.message.register(part_qty, PartAdd.waiting_for_qty)
    dp.message.register(staff_add_id, AddOperator.waiting_for_id)
    dp.message.register(partcode_add_code, AdminPartCodeAdd.waiting_for_code)
    dp.message.register(partcode_add_name, AdminPartCodeAdd.waiting_for_name)
    dp.message.register(partcode_add_photo, AdminPartCodeAdd.waiting_for_photo)
    dp.message.register(add_printer_ip, AdminAddPrinter.waiting_for_ip)
    dp.message.register(remove_printer_ip, AdminRemovePrinter.waiting_for_ip)
    dp.message.register(report_other_text, ReportOther.waiting_for_text)

    flask_thread = threading.Thread(target=run_flask, daemon=True)
    flask_thread.start()

    await bot.delete_webhook(drop_pending_updates=True)
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
